# endgame
import os
import sys
import lxml.html
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pytz


def fill_excel(results_lst, excel_filename):
    col = 0
    for key, value in results_lst.items():
        col += 1
        ws.cell(row=1, column=col).value = key
        max = len(key)
        for index, val in enumerate(value):
            if (key == "Date Posted") and val != '':
                try:
                    dt = datetime.strptime(val, '%Y-%m-%d %H:%M:%S.%f')
                except ValueError:
                    dt = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
                dt_timedelta = timedelta(hours=5, minutes=30)
                dt = dt + dt_timedelta
                ws.cell(row=index + 2,
                        column=col).value = dt.strftime('%d-%b, %I:%M:%S %p %Z')
            else:
                ws.cell(row=index + 2, column=col).value = val
            if len(val) > max:
                max = len(val)
            ws.column_dimensions[get_column_letter(col)].width = max + 3

    wb.save(excel_filename)


def parse_data_from_html(column_lst, filename):
    with open(filename, 'r') as file:
        html_content = str(file.read())
        doc = lxml.html.fromstring(html_content)

    result_lst = {}
    remove_index = []
    for data in column_list:
        fp = "//td[@data-title='" + data + "']/text()"
        lst = doc.xpath(fp)
        temp_lst = []
        for index, item in enumerate(lst):
            it = str(item).strip()
            temp_lst.append(it)

        new_tmp = []
        for i, j in enumerate(temp_lst):
            new_tmp.append(j)
        result_lst[data] = new_tmp

    return result_lst


if __name__ == '__main__':
    with open('config.json', 'r') as file:
        data = json.load(file)

    table_name = sys.argv[1].lower()
    if table_name == "user":
        html_filename = data["User_Html"]
        excel_filename = data["User_Excel"]
        column_list = data["User"]
    elif table_name == "post":
        html_filename = data["Post_Html"]
        excel_filename = data["Post_Excel"]
        column_list = data["Post"]

    filepath = os.path.join(os.getcwd(), excel_filename)
    if not os.path.isfile(filepath):
        os.system('cp ./EmptyExcelFiles/' + excel_filename + ' .')

    wb = load_workbook(filename=excel_filename)
    ws = wb.active

    results_lst = parse_data_from_html(column_list, html_filename)
    fill_excel(results_lst, excel_filename)
